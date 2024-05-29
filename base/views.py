from django.shortcuts import render, redirect
from django.shortcuts import render, get_object_or_404

from django.http import HttpResponse,HttpResponseRedirect
from django.contrib import messages
from .models import *
from .forms import BookForm,StudentForm,IssueForm


from django.contrib.auth.decorators import login_required




@login_required(login_url = '/admin_login')
def book_list(request):
    books = Book.objects.all()
    return render(request, 'base/book_list.html', {'books':books} )

@login_required(login_url = '/admin_login')
def student_list(request):
    students = Student.objects.all()
    return render(request, 'base/student_list.html', {'students':students} )



@login_required(login_url = '/admin_login')
def add_book(request):
    form = BookForm()
    if request.method == 'POST':
        form = BookForm(request.POST)
        if form.is_valid():
            book = form.save(commit=False)  # Create a book instance but don't save it yet
            
            # Perform additional validation
            isbn = form.cleaned_data['isbn']
            if len(isbn) != 10:
                form.add_error('isbn', 'ISBN must be exactly 10 characters long.')


            if book.total_copies < 0:
                form.add_error('total_copies', 'Total copies must be a non-negative number.')
            
            if book.available_copies < 0:
                form.add_error('available_copies', 'Available copies must be a non-negative number.')
            
            if book.available_copies > book.total_copies:
                form.add_error('available_copies', 'Available copies cannot exceed total copies.')

            if book.total_copies != book.available_copies:
                form.add_error('total_copies', 'Total copies must equal available copies.')

            if form.errors:
                # If there are any validation errors, render the form with errors
                context = {'form': form}
                return render(request, 'base/add_book.html', context)
            
            # All validation checks passed, save the book
            book.save()
            
            return redirect('base:book_list')

    context = {'form': form}
    return render(request, 'base/add_book.html', context)



@login_required(login_url = '/admin_login')
def update_book(request, pk):
    book = Book.objects.get(pk=pk)
    form = BookForm(instance=book)
    if request.method == 'POST':
        form = BookForm(request.POST, instance=book)
        if form.is_valid():
            updated_book = form.save(commit=False)  # Create a book instance but don't save it yet
            
            # Perform additional validation
            if updated_book.total_copies < 0:
                form.add_error('total_copies', 'Total copies must be a non-negative number.')
            
            isbn = form.cleaned_data['isbn']
            if len(isbn) != 10:
                form.add_error('isbn', 'ISBN must be exactly 10 characters long.')
            
            if updated_book.available_copies < 0:
                form.add_error('available_copies', 'Available copies must be a non-negative number.')
            
            if updated_book.available_copies > updated_book.total_copies:
                form.add_error('available_copies', 'Available copies cannot exceed total copies.')

            if form.errors:
                # If there are any validation errors, render the form with errors
                context = {'form': form}
                return render(request, 'base/add_book.html', context)
            
            # All validation checks passed, save the updated book
            updated_book.save()
            
            return redirect('base:book_list')

    context = {'form': form}
    return render(request, 'base/add_book.html', context)



@login_required(login_url = '/admin_login')
def delete_book(request, pk):
    book = get_object_or_404(Book, pk=pk)
    
    # Check if the book has been issued
    if Issue.objects.filter(book=book, status='issued').exists():
        return HttpResponse('Cannot delete book with has been issued.')
    # If the book has not been issued, proceed with the deletion.
    book.delete()
    return HttpResponseRedirect(reverse('base:book_list'))

# def delete_book(request, pk):
#     book = Book.objects.get(pk=pk)
#     book.delete()
#     return HttpResponseRedirect(reverse('base:book_list'))



@login_required(login_url = '/admin_login')
def update_student(request,pk):

    student = Student.objects.get(pk=pk)
    form = StudentForm(instance=student)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance = student)
        if form.is_valid():
            form.save()
            return redirect('base:student_list')

    context = {'form':form}
    return render(request,'base/student_update.html', context)







@login_required(login_url = '/admin_login')
def delete_student(request, pk):
    student = get_object_or_404(Student, pk=pk) 
    # Check if the student has any issued books
    if Issue.objects.filter(student=student, status='issued').exists():
        return HttpResponse('Cannot delete student with issued books.')
    student.delete()
    return HttpResponseRedirect(reverse('base:student_list'))






#user chai create vayo and also student done ##############################
from django.shortcuts import render, redirect
from django.urls import reverse
from openpyxl import load_workbook
from django.contrib.auth.models import User

@login_required(login_url = '/admin_login')
def upload(request):
    if request.method == 'POST' and request.FILES['myfile']:
        myfile = request.FILES['myfile']
        # Check that the file extension is .xlsx
        if not myfile.name.endswith('.xlsx'):
            return render(request, 'base/student_import.html', {'error': 'File must be in .xlsx format.'})
        wb = load_workbook(filename=myfile, read_only=True)
        ws = wb.active 
        for row in ws.iter_rows(min_row=2):
            username = row[2].value  # Assuming email is used as the username
            password = str(row[3].value)  # Convert password to string
            email = row[2].value
            user = User.objects.create_user(username=username, password=password, email=email)
            # Additional fields specific to your application can be added here
            user.save()
            student = Student(
                student_id=row[0].value,
                name=row[1].value,
                email=row[2].value,
                password=row[3].value,
                course=row[4].value,
                batch=row[5].value,
                phone_number=row[6].value
            )
            student.save()
        return redirect(reverse('base:student_list'))
    return render(request, 'base/student_import.html')









from datetime import date

@login_required(login_url = '/admin_login')
def issue_detail(request):
    today = date.today()
    issues = Issue.objects.all()

    for issue in issues:
        if issue.expiry_date < today and issue.status == 'issued':
            days_overdue = (today - issue.expiry_date).days
            fine = days_overdue * 10  

            # # Limit the fine to a maximum of 500
            # if fine > 500:
            #     fine = 500
            issue.fine = fine
            issue.save()
    return render(request, 'base/issue/issue_detail.html', {'issues': issues})






#perfect for my project
@login_required(login_url = '/admin_login')
def issue_book(request):
    form = IssueForm()
    if request.method == 'POST':
        form = IssueForm(request.POST)
        if form.is_valid():
            book = form.cleaned_data['book']
            student = form.cleaned_data['student']

            # Check if the book has already been issued to the same user
            if Issue.objects.filter(book=book, student=student, status='issued').exists():
                form.add_error('book', 'This book has already been issued to the same user.')  #book field ma error halxa
            
            else:
                if book.available_copies > 0:
                    book.available_copies -= 1
                    book.save()

                    form.save()
                    return redirect('base:issue_detail')
                else:
                    form.add_error('book', 'This book is currently not available.')

    context = {'form': form}
    return render(request, 'base/issue/issue_book.html', context)


@login_required(login_url = '/admin_login')
def return_book(request, issue_id):
    issue = Issue.objects.get(id=issue_id)
    if issue.status == 'issued':
        issue.status = 'returned'
        issue.fine = 0
        issue.save()

        book = issue.book
        book.available_copies += 1
        book.save()

        messages.success(request, 'Book returned successfully.')
    else:
        messages.error(request, 'Book is already returned')

    return redirect('base:issue_detail')



from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .models import Student

@login_required(login_url = '/student_login')
def profile(request):
    user = request.user  # Retrieve the logged-in user
    try:
        student = Student.objects.get(email=user.email)  # Retrieve the student using the email field
        context = {'student': student}  
        return render(request, 'base/student/profile.html', context)
    except Student.DoesNotExist:
        return render(request, "base/student/profile.html", context)



@login_required(login_url = '/student_login')
def change_password(request):
    if request.method == "POST":
        current_password = request.POST['current_password']
        new_password = request.POST['new_password']
        try:
            u = User.objects.get(id=request.user.id)
            if u.check_password(current_password):
                u.set_password(new_password)
                u.save()
                alert = True
                return render(request, "base/student/change_password.html", {'alert':alert})
            else:
                currpasswrong = True
                return render(request, "base/student/change_password.html", {'currpasswrong':currpasswrong})
        except:
            pass
    return render(request, "base/student/change_password.html")









@login_required(login_url = '/student_login')
def view_student_book(request):
    student = Student.objects.get(email=request.user.email)  # Retrieve the student using the email field
    today = date.today()
    issues = Issue.objects.filter(student=student, status='issued')  # Filter issues by student and status

    for issue in issues:
        if issue.expiry_date < today:
            days_overdue = (today - issue.expiry_date).days
            fine = days_overdue * 10
            # if fine > 500:
            #     fine = 500
            issue.fine = fine
            issue.save()

    return render(request, 'base/student/viewbook.html', {'issues': issues})






from django.contrib.auth import authenticate, login, logout


def home_view(request):
    return render(request,'base/index.html')


def admin_login(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)

        if user is not None:
            login(request, user)
            if request.user.is_superuser:
                return redirect("base:book_list")
            else:
                return HttpResponse("You are not an admin.")
        else:
            alert = True
            return render(request, "base/accounts/adminlogin.html", {'alert':alert})
    return render(request, "base/accounts/adminlogin.html")



def student_login(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)

        if user is not None:
            login(request, user)
            if request.user.is_superuser:
                return HttpResponse("You are not a student!!")
            else:
                return redirect("base:profile")
        else:
            alert = True
            return render(request, "base/accounts/student_login.html", {'alert':alert})
    return render(request, "base/accounts/student_login.html")




def Logout_User(request):
    logout(request)
    return redirect ("base:index")




######
######
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.http import HttpResponseRedirect
from .models import Book, Issue, Student

def request_book(request, book_id):
    book = get_object_or_404(Book, pk=book_id)
    student = get_object_or_404(Student, student_id=request.user.username)  # Assuming you store the student ID in the username field

    # Check if the book is available
    if book.available_copies > 0:
        # Create a new issue record
        issue = Issue(student=student, book=book, status='requested')
        issue.save()

        # Decrease the available copies of the book
        book.available_copies -= 1
        book.save()

        return HttpResponseRedirect(reverse('base:book_list'))  # Redirect to the book list page or a confirmation page
    else:
        return render(request, 'base/error_page.html', {'message': 'Sorry, the book is not available for request.'})
